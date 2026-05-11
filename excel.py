import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

STATUS_LABELS = {
    "pendente":  "Aguarda aprovação",
    "aprovada":  "Aprovada",
    "recusada":  "Recusada",
    "concluida": "Concluída",
}

def _header(ws, row, n_cols, bg="1F5AA8"):
    fill = PatternFill("solid", fgColor=bg)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def gerar_historico(negs: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Negociações"
    thin = Border(bottom=Side(style="thin", color="EEEEEE"))

    headers = ["ID", "Data", "Fornecedor", "CNPJ", "NF(s)", "Vencimentos",
               "Desdobramentos", "Valor Total", "Taxa (%)", "Ganho",
               "Valor Pago", "Status", "Criado por", "Aprovador", "Decisão em", "Obs"]
    ws.append(headers)
    _header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24

    for n in negs:
        aprov = n.get("aprovador_id") or "Não requerido"
        decisao = ""
        if n.get("decisao_em"):
            try:
                decisao = datetime.fromisoformat(n["decisao_em"]).strftime("%d/%m/%Y")
            except Exception:
                decisao = n["decisao_em"][:10]

        notas = n["notas"] if isinstance(n["notas"], list) else []
        ws.append([
            n["id"],
            datetime.fromisoformat(n["criado_em"]).strftime("%d/%m/%Y"),
            n["fornecedor"],
            n.get("cnpj", ""),
            "; ".join(x["nf"] for x in notas),
            "; ".join(x.get("vencimento", "") for x in notas),
            "; ".join(x.get("desdobramento", "") for x in notas),
            float(n["valor_total"]),
            float(n["taxa"]),
            float(n["ganho"]),
            float(n["valor_pago"]),
            STATUS_LABELS.get(n["status"], n["status"]),
            n.get("criado_por", ""),
            aprov,
            decisao,
            n.get("obs", ""),
        ])
        r = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).border = thin
        for c in [8, 10, 11]:
            ws.cell(r, c).number_format = 'R$ #,##0.00'
        ws.cell(r, 9).number_format = '0.00"%"'
        status_colors = {"concluida": "E1F5EE", "aprovada": "E1F5EE",
                         "pendente": "FAEEDA", "recusada": "FCEBEB"}
        if n["status"] in status_colors:
            ws.cell(r, 12).fill = PatternFill("solid", fgColor=status_colors[n["status"]])

    _set_col_widths(ws, [12,12,24,18,20,20,18,16,10,14,16,18,18,22,14,36])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_relatorio(concluidas: list, periodo: str) -> bytes:
    ganho_total = sum(float(n["ganho"]) for n in concluidas)
    taxa_media = (sum(float(n["taxa"]) for n in concluidas) / len(concluidas)) if concluidas else 0
    vol_total = sum(float(n["valor_total"]) for n in concluidas)

    wb = openpyxl.Workbook()

    # Resumo
    ws1 = wb.active
    ws1.title = "Resumo Executivo"
    ws1["A1"] = f"Relatório de Factoring — {periodo}"
    ws1["A1"].font = Font(bold=True, size=14, color="1F5AA8")
    ws1["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="888888", size=10)

    dados = [
        ("", ""),
        ("RESUMO EXECUTIVO", ""),
        ("Negociações concluídas", len(concluidas)),
        ("Ganho total (R$)", ganho_total),
        ("Taxa média negociada (%)", taxa_media),
        ("Volume total negociado (R$)", vol_total),
    ]
    for i, (lab, val) in enumerate(dados, 4):
        ws1.cell(i, 1, lab).font = Font(bold=True, size=11)
        cell = ws1.cell(i, 2, val)
        if "R$" in str(lab): cell.number_format = 'R$ #,##0.00'
        elif "%" in str(lab): cell.number_format = '0.00"%"'
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 22

    # Detalhamento
    ws2 = wb.create_sheet("Detalhamento")
    headers2 = ["Fornecedor", "NF(s)", "Valor (R$)", "Taxa (%)", "Ganho (R$)", "Valor Pago (R$)", "Data", "Aprovador"]
    ws2.append(headers2)
    _header(ws2, 1, len(headers2))
    ws2.freeze_panes = "A2"
    thin = Border(bottom=Side(style="thin", color="EEEEEE"))

    for n in sorted(concluidas, key=lambda x: x["criado_em"], reverse=True):
        notas = n["notas"] if isinstance(n["notas"], list) else []
        ws2.append([
            n["fornecedor"],
            "; ".join(x["nf"] for x in notas),
            float(n["valor_total"]),
            float(n["taxa"]),
            float(n["ganho"]),
            float(n["valor_pago"]),
            datetime.fromisoformat(n["criado_em"]).strftime("%d/%m/%Y"),
            n.get("aprovador_id") or "—",
        ])
        r = ws2.max_row
        for c in [3, 5, 6]: ws2.cell(r, c).number_format = 'R$ #,##0.00'
        ws2.cell(r, 4).number_format = '0.00"%"'
        for c in range(1, len(headers2)+1): ws2.cell(r, c).border = thin

    last = ws2.max_row
    ws2.append(["TOTAL","", f"=SUM(C2:C{last})","", f"=SUM(E2:E{last})", f"=SUM(F2:F{last})","",""])
    r = ws2.max_row
    for c in range(1, 9): ws2.cell(r, c).font = Font(bold=True)
    for c in [3, 5, 6]: ws2.cell(r, c).number_format = 'R$ #,##0.00'
    _set_col_widths(ws2, [26,22,16,10,14,16,12,22])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
